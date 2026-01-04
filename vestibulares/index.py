import tabula
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- ARQUIVOS PDF ----------
pdf_unesp_basico = "/mnt/data/Unesp Ciências Biológicas base (2 primeiros anos).pdf"
pdf_unesp_bach = "/mnt/data/Unesp Ciências Biológicas Bacharelado (2).pdf"
pdf_famerp = "/mnt/data/Estrutura-Curricular-Medicina-2023.pdf"

# ---------- FUNÇÃO PARA LER TABELAS DOS PDFs ----------
def ler_pdf(pdf_path):
    try:
        tabelas = tabula.read_pdf(pdf_path, pages="all", multiple_tables=True, lattice=True)
        return tabelas
    except Exception as e:
        print("Erro ao ler PDF:", pdf_path, e)
        return []

# ---------- LER TODOS OS PDF ----------
tabelas_unesp_basico = ler_pdf(pdf_unesp_basico)
tabelas_unesp_bach = ler_pdf(pdf_unesp_bach)
tabelas_famerp = ler_pdf(pdf_famerp)

# ---------- FUNÇÃO PARA ORGANIZAR AS TABELAS ----------
def organizar_dataframe(lista_tabelas, curso, instituicao):
    dfs = []
    for tabela in lista_tabelas:
        df = tabela.copy()
        df.columns = [str(col) for col in df.columns]
        df.insert(0, "Curso", curso)
        df.insert(1, "Instituição", instituicao)
        dfs.append(df)
    if dfs:
        return pd.concat(dfs, ignore_index=True)
    return pd.DataFrame()

# Criar dataframes organizados
df_unesp_basico = organizar_dataframe(tabelas_unesp_basico, "Ciências Biológicas", "UNESP")
df_unesp_bach = organizar_dataframe(tabelas_unesp_bach, "Ciências Biológicas (Bacharelado)", "UNESP")
df_famerp = organizar_dataframe(tabelas_famerp, "Medicina", "FAMERP")

# ---------- CRIAR ARQUIVO EXCEL ----------
wb = Workbook()

# Remover aba padrão
default_sheet = wb.active
wb.remove(default_sheet)

# ---------- FUNÇÃO PARA ADICIONAR ABA COM ESTILO ----------
def adicionar_aba_formatada(wb, df, nome_aba):
    ws = wb.create_sheet(title=nome_aba)

    # Cores
    verde_cabecalho = "98FB98"  # verde claro
    verde_linha = "E6FFE6"      # tom verde bem suave
    branco = "FFFFFF"

    # Escrever DataFrame na aba
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)

    # Formatar cabeçalho
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=verde_cabecalho, fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Linhas alternadas
    for row in ws.iter_rows(min_row=2):
        r = row[0].row
        fill_color = verde_linha if r % 2 == 0 else branco
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color, fill_type="solid")

    # Auto-ajuste da largura das colunas
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

# ---------- ADICIONAR ABAS ----------
if not df_unesp_basico.empty:
    adicionar_aba_formatada(wb, df_unesp_basico, "Unesp - Básico")

if not df_unesp_bach.empty:
    adicionar_aba_formatada(wb, df_unesp_bach, "Unesp - Bacharelado")

if not df_famerp.empty:
    adicionar_aba_formatada(wb, df_famerp, "Medicina - FAMERP")

# ---------- SALVAR ----------
arquivo_final = "talles_disciplinas.xlsx"
wb.save(arquivo_final)

print("Planilha criada com sucesso:", arquivo_final)